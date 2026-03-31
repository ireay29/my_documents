#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
日本の運送・配送会社リスト作成スクリプト
観点：IT未導入・誤配送・取り間違い課題を抱えている会社
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ===== データ定義 =====
companies = [
    # --- 確度高（IT未導入の直接証拠 or 誤配送の直接証拠あり） ---
    {
        "no": 1,
        "company_name": "コスモ運輸株式会社",
        "location": "埼玉県戸田市笹目北町2-24（東京都文京区に本社登記）",
        "region": "関東",
        "business_type": "一般貨物自動車運送（製本・建材・家具等）",
        "employee_size": "約60〜70名",
        "vehicle_count": "大型11台・4t9台・2t32台（計52台）",
        "website": "https://www.cosmounyu.co.jp/",
        "it_lag_score": 5,
        "delivery_error_score": 4,
        "total_score": 9,
        "it_lag_evidence": "2024年末まで紙の配車表を手書きで作成し、FAXでやり取りする日々が続いていた（アセンド・ロジックス導入事例より）",
        "delivery_error_evidence": "手書き配車表・FAX管理による情報連携の遅延・ミスリスクが高い状態だった",
        "source": "https://note.ascendlogi.co.jp/n/nbd7d648ecd3e",
        "notes": "2024年末にロジックスを導入し改善中。導入前の状態が典型的なIT未導入事例として公開されている。",
        "priority": "高",
    },
    {
        "no": 2,
        "company_name": "荻布倉庫株式会社",
        "location": "富山県高岡市荻布字川開688番地",
        "region": "北陸・中部",
        "business_type": "一般貨物自動車運送・倉庫業",
        "employee_size": "約91名",
        "vehicle_count": "不明",
        "website": "https://oginosoko.com/",
        "it_lag_score": 5,
        "delivery_error_score": 3,
        "total_score": 8,
        "it_lag_evidence": "IT技術担当者が社内に1名しかおらず、配車表のデジタル化を複数のITベンダーに断られていた。ロジックス導入前は配車担当者が深夜まで残業する状態（アセンド事例より）",
        "delivery_error_evidence": "配車表が手作業管理のため、情報の抜け漏れ・ミスが発生しやすい状態だった",
        "source": "https://www.logi-square.com/column/interview/221005",
        "notes": "明治30年（1897年）創業の老舗。ロジックス導入後は配車担当の残業が大幅削減。",
        "priority": "高",
    },
    {
        "no": 3,
        "company_name": "九州トランスポート株式会社（旧・有限会社日向商運）",
        "location": "宮崎県",
        "region": "九州",
        "business_type": "一般貨物自動車運送（原乳・タイヤ・肥料・医薬品等）",
        "employee_size": "約37名",
        "vehicle_count": "不明",
        "website": "不明",
        "it_lag_score": 5,
        "delivery_error_score": 4,
        "total_score": 9,
        "it_lag_evidence": "「以前は経理も配車も紙にボールペンという状態だった」「昔ながらのやり方でずっときた会社。トラックの管理や時間管理もドライバーに任せているような状態」（M&A事例記事より）",
        "delivery_error_evidence": "管理体制がドライバー任せで、配送ミス・取り間違いのリスクが高い状態だった",
        "source": "https://ma-succeed.jp/content/agreement/post-13358",
        "notes": "M&Aによりフジグループ入り後、運行管理システムを導入。導入前の状態が典型的なアナログ運送会社の事例。",
        "priority": "高",
    },
    {
        "no": 4,
        "company_name": "株式会社T.M.G",
        "location": "大阪府茨木市",
        "region": "関西",
        "business_type": "貨物軽自動車運送業・小口配送事業（Amazonデリバリープロバイダ）",
        "employee_size": "230名（グループ登録台数2,200台）",
        "vehicle_count": "登録台数2,200台（グループ全体）",
        "website": "https://tmg-group.co.jp/",
        "it_lag_score": 3,
        "delivery_error_score": 5,
        "total_score": 8,
        "it_lag_evidence": "Amazonデリバリープロバイダとして急拡大した結果、配送品質管理が追いついていない状態",
        "delivery_error_evidence": "Amazonセラーセントラルフォーラムに「不在票なし」「ポストに無理やり詰め込む」「配達時間指定無視」「荷物の紛失・破損」など具体的な誤配送・配送品質トラブルの報告が多数",
        "source": "https://sellercentral.amazon.co.jp/seller-forums/discussions/t/df4f1d782c476484ce3c2f0c22e9c728",
        "notes": "関西を中心に西日本に多くの拠点を持つ。口コミサイトやフォーラムで誤配送や配送品質に関する不満が多い。",
        "priority": "高",
    },
    {
        "no": 5,
        "company_name": "東京運送株式会社",
        "location": "東京都江東区亀戸7-39-25",
        "region": "関東",
        "business_type": "石油製品輸送（潤滑油タンク・ローリー）",
        "employee_size": "約14〜25名",
        "vehicle_count": "大型ローリー車11台・その他6台（計17台）",
        "website": "http://tokyounso.co.jp/",
        "it_lag_score": 5,
        "delivery_error_score": 3,
        "total_score": 8,
        "it_lag_evidence": "「それまで全て手書き・手計算だったアナログ業務」（アセンド事例より）。M&A後にロジックスを導入して改善。",
        "delivery_error_evidence": "手書き・手計算管理による情報の正確性リスクが存在していた",
        "source": "https://note.ascendlogi.co.jp/m/mf5a20e647758",
        "notes": "北海道の石油輸送株式会社にM&Aされたことをきっかけにアナログ業務改革に踏み切った。",
        "priority": "高",
    },
    {
        "no": 6,
        "company_name": "株式会社朋友",
        "location": "大阪府堺市中区平井848番",
        "region": "関西",
        "business_type": "一般貨物運送事業・貨物運送取扱事業（関東〜関西間幹線）",
        "employee_size": "約117〜155名",
        "vehicle_count": "約150台",
        "website": "https://foryou1992.co.jp/",
        "it_lag_score": 4,
        "delivery_error_score": 3,
        "total_score": 7,
        "it_lag_evidence": "労基署からの是正勧告をきっかけにDXを推進。ロジックス導入前は労務管理が属人化・アナログ管理だった（アセンド事例より）",
        "delivery_error_evidence": "業務の属人化・アナログ管理による情報連携ミスのリスクが存在",
        "source": "https://logix.ascendlogi.co.jp/case/foryou",
        "notes": "1992年創業。是正勧告を機にDXを推進し、3時間の集計作業を10分に短縮。",
        "priority": "中",
    },
    {
        "no": 7,
        "company_name": "會津通運株式会社",
        "location": "福島県会津若松市町北町大字始字見島83",
        "region": "東北",
        "business_type": "一般貨物自動車運送・JRコンテナ輸送・倉庫業",
        "employee_size": "約86〜90名",
        "vehicle_count": "不明",
        "website": "https://aizu-tsuuun.co.jp/",
        "it_lag_score": 4,
        "delivery_error_score": 3,
        "total_score": 7,
        "it_lag_evidence": "ロジックス導入前は紙ベースの管理が中心。「常に新しい挑戦を求める社長のもと」ITツール導入を決断（アセンド事例より）",
        "delivery_error_evidence": "紙ベース管理による情報の抜け漏れリスクが存在していた",
        "source": "https://note.ascendlogi.co.jp/m/mf5a20e647758",
        "notes": "昭和25年（1950年）創業。資本金2,000万円。福島県内に2営業所。",
        "priority": "中",
    },
    {
        "no": 8,
        "company_name": "大三ロジテック株式会社",
        "location": "長野県上伊那郡箕輪町中箕輪6597",
        "region": "北陸・中部",
        "business_type": "一般貨物自動車運送・精密機器輸送・オフィス移転",
        "employee_size": "約23〜31名",
        "vehicle_count": "不明",
        "website": "https://www.daisanlogi.co.jp/",
        "it_lag_score": 4,
        "delivery_error_score": 3,
        "total_score": 7,
        "it_lag_evidence": "1年半前に他社クラウドサービスを導入したが活用終了。ロジックス導入前は受注・配車情報の可視化ができていなかった（アセンド事例より）",
        "delivery_error_evidence": "受注・配車情報の見える化が不十分で、ミス・抜け漏れのリスクがあった",
        "source": "https://note.ascendlogi.co.jp/m/mf5a20e647758",
        "notes": "1964年創業。精密機器輸送に特化。神奈川トラック協会青年部委員長も務める。",
        "priority": "中",
    },
    {
        "no": 9,
        "company_name": "山田運送株式会社",
        "location": "大阪府東大阪市",
        "region": "関西",
        "business_type": "一般貨物自動車運送・幹線輸送・地域配送",
        "employee_size": "約268名",
        "vehicle_count": "不明",
        "website": "https://www.yamada-transport.co.jp/",
        "it_lag_score": 4,
        "delivery_error_score": 2,
        "total_score": 6,
        "it_lag_evidence": "エン転職の口コミに「業界的にアナログ」「紙の伝票やFAXでのやりとりが多い」との記述あり",
        "delivery_error_evidence": "直接的な誤配送の記述は見当たらないが、アナログ管理体制からリスクが推測される",
        "source": "https://employment.en-japan.com/comp-107002/",
        "notes": "大阪を中心に関東〜関西間の幹線輸送を展開。口コミからアナログな業務体制がうかがえる。",
        "priority": "中",
    },
    {
        "no": 10,
        "company_name": "株式会社ヒサノ",
        "location": "熊本県熊本市南区南高江2-1-15",
        "region": "九州",
        "business_type": "半導体製造装置輸送・精密機器輸送・倉庫業",
        "employee_size": "約68〜84名",
        "vehicle_count": "不明",
        "website": "https://www.kk-hisano.co.jp/",
        "it_lag_score": 4,
        "delivery_error_score": 3,
        "total_score": 7,
        "it_lag_evidence": "改革当初は「情報関連部署もなければデジタル人材もいない、いわゆるアナログ企業だった」。配車業務は1人の担当者が紙媒体で管理（神戸DXお助け隊事例より）",
        "delivery_error_evidence": "配車担当への業務集中・紙管理による情報連携ミスが発生。2016年熊本地震後の需要急増時に問題が顕在化",
        "source": "https://kobe-dxotasuketai.jp/jirei/page10.html",
        "notes": "1935年創業。2021年にDX認定事業者に選出。現在はDX化が進んでいるが、導入前の課題が公開されている。",
        "priority": "中",
    },
    {
        "no": 11,
        "company_name": "永山運送株式会社",
        "location": "東京都多摩市（本社）、神奈川・埼玉等に営業所",
        "region": "関東",
        "business_type": "一般貨物自動車運送（冷凍・冷蔵・常温）",
        "employee_size": "約261〜286名",
        "vehicle_count": "冷凍車177台・常温車16台（計193台）",
        "website": "https://www.nagayamaunsou.co.jp/",
        "it_lag_score": 4,
        "delivery_error_score": 3,
        "total_score": 7,
        "it_lag_evidence": "「配車計画は人依存で膨大な作業量だった」。Tradiss導入前は全5拠点で人手による配車管理（Tradiss導入事例より）",
        "delivery_error_evidence": "人依存の配車管理による情報の抜け漏れ・ミスリスクが存在していた",
        "source": "https://tradiss-web.com/casestudy/detail.php?eid=00005",
        "notes": "首都圏区域免許。冷凍・冷蔵輸送を主力とする。Tradiss導入後は全5拠点でデジタル化を推進中。",
        "priority": "中",
    },
    {
        "no": 12,
        "company_name": "カホク運送株式会社",
        "location": "宮城県仙台市宮城野区中野1丁目2-15",
        "region": "東北",
        "business_type": "一般貨物自動車運送（東北〜関西間）",
        "employee_size": "約40名",
        "vehicle_count": "不明",
        "website": "https://www.kahoku-sendai.com/",
        "it_lag_score": 3,
        "delivery_error_score": 3,
        "total_score": 6,
        "it_lag_evidence": "中小企業基盤整備機構のIT導入支援事例として掲載。DX化された運行管理システムを導入済みだが、導入前はアナログ管理だった",
        "delivery_error_evidence": "IT導入前は配送情報の管理が不十分で、ミスのリスクが存在していた",
        "source": "https://digiwith.smrj.go.jp/cocoapp/case/cp577f0000004h9c.html",
        "notes": "1992年設立。資本金1,000万円。東北から関西まで広域輸送を展開。現在はDX化が進んでいる。",
        "priority": "中",
    },
    {
        "no": 13,
        "company_name": "南国運送有限会社",
        "location": "高知県南国市小籠752番地1",
        "region": "四国",
        "business_type": "一般貨物自動車運送",
        "employee_size": "不明（車両台数19台の小規模）",
        "vehicle_count": "19台",
        "website": "https://nangoku-unsou.jp/",
        "it_lag_score": 4,
        "delivery_error_score": 3,
        "total_score": 7,
        "it_lag_evidence": "全日本トラック協会の物流DX戦略セミナー資料に「データ経営実践事例」として掲載。「1959年の創業以来、中型車に特化」した地方の小規模運送会社でDX推進中",
        "delivery_error_evidence": "小規模・地方の運送会社として、IT管理の遅れによる配送ミスリスクが推測される",
        "source": "https://www.gta.or.jp/maintenance/upload/file/seinen/download/dxseminasiryo.pdf",
        "notes": "1959年創業。高知県南国市に本社・倉庫を構える地方の小規模運送会社。",
        "priority": "中",
    },
    {
        "no": 14,
        "company_name": "小倉貨物運輸株式会社",
        "location": "福岡県北九州市小倉北区西港町84-8",
        "region": "九州",
        "business_type": "一般区域貨物運輸・商品保管発送業務",
        "employee_size": "約30名",
        "vehicle_count": "27台",
        "website": "https://k-unyu.co.jp/",
        "it_lag_score": 4,
        "delivery_error_score": 3,
        "total_score": 7,
        "it_lag_evidence": "北九州市の小規模運送会社。ウェブサイトにIT活用の記述がなく、アナログ管理の可能性が高い",
        "delivery_error_evidence": "5t平ボデー・ゲート車など多様な車両を少人数で管理しており、配送ミスのリスクが推測される",
        "source": "https://k-unyu.co.jp/about-us/overview.html",
        "notes": "北九州市小倉北区の地場運送会社。一般区域貨物・路線貨物・商品保管発送を手がける。",
        "priority": "中",
    },
    {
        "no": 15,
        "company_name": "株式会社MPJ",
        "location": "愛知県名古屋市港区",
        "region": "東海",
        "business_type": "一般貨物自動車運送・貨物利用運送・倉庫業",
        "employee_size": "約10名",
        "vehicle_count": "不明",
        "website": "不明",
        "it_lag_score": 5,
        "delivery_error_score": 3,
        "total_score": 8,
        "it_lag_evidence": "求人票に「手書きの伝票」と明記されており、アナログ管理が現在も継続中",
        "delivery_error_evidence": "手書き伝票管理による誤記・取り間違いのリスクが高い",
        "source": "https://xn--pckua2a7gp15o89zb.com/jbi/75778d40dc260f0aacc9ef7f7853774a",
        "notes": "名古屋港区の小規模運送会社。求人票から手書き伝票が現在も使用されていることが確認できた。",
        "priority": "高",
    },
    {
        "no": 16,
        "company_name": "株式会社セイリョウ",
        "location": "埼玉県戸田市",
        "region": "関東",
        "business_type": "運送事業（家具・精密機器・食品・引越）・介護事業（福祉用具・介護タクシー）",
        "employee_size": "約42名（パートナースタッフ含む）",
        "vehicle_count": "不明",
        "website": "不明",
        "it_lag_score": 4,
        "delivery_error_score": 3,
        "total_score": 7,
        "it_lag_evidence": "Instagramに「手書き日報作成 ❌ 月末の伝票整理」と記載。勤怠管理もアナログ管理だった",
        "delivery_error_evidence": "配送スピード低下・収入減の課題に言及。アナログ管理による配送ミスリスクが推測される",
        "source": "https://www.instagram.com/p/DUPKu2nEotw/",
        "notes": "戸田市の中小運送会社。運送と介護を兼業。2024年問題への対応として給与制度見直しを実施。",
        "priority": "中",
    },
    {
        "no": 17,
        "company_name": "フレッシュ物流株式会社",
        "location": "三重県松阪市",
        "region": "東海",
        "business_type": "食品配送（冷凍・冷蔵・常温）",
        "employee_size": "不明（中小規模）",
        "vehicle_count": "不明",
        "website": "https://fresh-logi.net/",
        "it_lag_score": 3,
        "delivery_error_score": 4,
        "total_score": 7,
        "it_lag_evidence": "「伝票と商品のチェックは必要です」という記述から、伝票によるアナログな商品チェックが現在も行われている",
        "delivery_error_evidence": "伝票と商品の照合が手作業であり、取り間違い・誤配送のリスクが高い。スーパー・飲食店・製造メーカー向けの多品種配送",
        "source": "https://fresh-logi.net/job/-/info/detail/M0000030",
        "notes": "三重県松阪市の食品配送会社。スーパー・飲食店・製造メーカー向けに冷凍・冷蔵・常温の食品を配送。",
        "priority": "中",
    },
    {
        "no": 18,
        "company_name": "芳誠流通株式会社",
        "location": "東京都大田区東海3-6-3",
        "region": "関東",
        "business_type": "青果物配送（市場→スーパー・集配センター等）",
        "employee_size": "約262名",
        "vehicle_count": "約180台",
        "website": "https://hr-hosei.co.jp/",
        "it_lag_score": 4,
        "delivery_error_score": 4,
        "total_score": 8,
        "it_lag_evidence": "「1979年の創業以来、アナログな業務に頼ってきた」。自社でのシステム開発を試みたが理想のDX実現に至らず、2024年にロジックスを導入（アセンド事例より）",
        "delivery_error_evidence": "青果物の多品種・多拠点配送でアナログ管理を継続。取り間違い・誤配送のリスクが高い業態",
        "source": "https://note.ascendlogi.co.jp/m/mf5a20e647758",
        "notes": "首都圏全域で青果物を配送。市場からスーパーや集配センターへのルート配送を担う。2024年にロジックス導入。",
        "priority": "高",
    },
    {
        "no": 19,
        "company_name": "株式会社西運",
        "location": "愛知県名古屋市中川区江松5-106",
        "region": "東海",
        "business_type": "全国配送一般貨物自動車運送",
        "employee_size": "約65〜70名",
        "vehicle_count": "60台",
        "website": "https://www.nishiun.com/",
        "it_lag_score": 3,
        "delivery_error_score": 3,
        "total_score": 6,
        "it_lag_evidence": "数年前に他社クラウドサービスを導入したが、サービス終了に伴い2023年にロジックスへ切り替え。導入前は紙の配車表を使用（アセンド事例より）",
        "delivery_error_evidence": "紙の配車表管理による情報連携ミスのリスクが存在していた",
        "source": "https://note.ascendlogi.co.jp/m/mf5a20e647758",
        "notes": "2011年設立。愛知・岐阜・三重・静岡の東海4県を中心に全国配送を展開。現在はロジックスで管理中。",
        "priority": "中",
    },
    {
        "no": 20,
        "company_name": "株式会社柳川合同",
        "location": "福岡県柳川市西浜武475-2",
        "region": "九州",
        "business_type": "一般貨物自動車運送・倉庫業",
        "employee_size": "約324〜416名（グループ計）",
        "vehicle_count": "不明",
        "website": "https://ygu.co.jp/",
        "it_lag_score": 3,
        "delivery_error_score": 3,
        "total_score": 6,
        "it_lag_evidence": "1954年創業の老舗地方運送会社。ウェブサイトにIT活用の具体的な記述がなく、地方の伝統的な運送会社としてIT化が遅れている可能性が高い",
        "delivery_error_evidence": "グループ全体で多数の車両・拠点を管理しており、アナログ管理による配送ミスリスクが推測される",
        "source": "https://ygu.co.jp/company.html",
        "notes": "1954年（昭和29年）設立。資本金6,900万円。福岡県柳川市の地場運送会社。",
        "priority": "中",
    },
]

# ===== Excel作成 =====
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "運送会社リスト"

# --- カラー定義 ---
COLOR_HEADER_BG = "1F3864"   # 濃紺
COLOR_HEADER_FG = "FFFFFF"   # 白
COLOR_HIGH_BG   = "FFE0E0"   # 薄赤（優先度：高）
COLOR_MID_BG    = "FFF3CD"   # 薄黄（優先度：中）
COLOR_ROW_ALT   = "F8F9FA"   # 薄グレー（交互行）
COLOR_SCORE_5   = "C6EFCE"   # 濃緑
COLOR_SCORE_4   = "FFEB9C"   # 黄
COLOR_SCORE_3   = "FFCC99"   # 橙
COLOR_SCORE_2   = "FFC7CE"   # 赤
COLOR_SCORE_1   = "FF9999"   # 濃赤

def score_color(score):
    if score >= 5: return COLOR_SCORE_5
    elif score == 4: return COLOR_SCORE_4
    elif score == 3: return COLOR_SCORE_3
    elif score == 2: return COLOR_SCORE_2
    else: return COLOR_SCORE_1

# --- ヘッダー定義 ---
headers = [
    ("No.", 5),
    ("会社名", 28),
    ("所在地", 30),
    ("地域", 8),
    ("業種・事業内容", 30),
    ("従業員数", 12),
    ("車両台数", 12),
    ("IT未導入\nスコア\n(1-5)", 10),
    ("誤配送\nリスク\nスコア(1-5)", 10),
    ("総合\nスコア", 8),
    ("優先度", 8),
    ("IT未導入の証拠・根拠", 50),
    ("誤配送・取り間違いの証拠・根拠", 50),
    ("情報ソース", 40),
    ("備考", 40),
    ("公式サイト", 35),
]

# --- ヘッダー行書き込み ---
header_font = Font(name="Meiryo UI", bold=True, color=COLOR_HEADER_FG, size=9)
header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

for col_idx, (header_text, col_width) in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header_text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = col_width

ws.row_dimensions[1].height = 45

# --- データ行書き込み ---
data_font = Font(name="Meiryo UI", size=9)
data_align_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
data_align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

for row_idx, c in enumerate(companies, start=2):
    priority = c["priority"]
    if priority == "高":
        row_bg = COLOR_HIGH_BG
    elif priority == "中":
        row_bg = COLOR_MID_BG
    else:
        row_bg = COLOR_ROW_ALT

    row_fill = PatternFill("solid", fgColor=row_bg)

    row_data = [
        c["no"],
        c["company_name"],
        c["location"],
        c["region"],
        c["business_type"],
        c["employee_size"],
        c["vehicle_count"],
        c["it_lag_score"],
        c["delivery_error_score"],
        c["total_score"],
        c["priority"],
        c["it_lag_evidence"],
        c["delivery_error_evidence"],
        c["source"],
        c["notes"],
        c["website"],
    ]

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.border = thin_border

        # スコア列は中央揃え＋スコアカラー
        if col_idx in (8, 9, 10):
            cell.alignment = data_align_center
            cell.fill = PatternFill("solid", fgColor=score_color(value))
        elif col_idx in (1, 4, 11):
            cell.alignment = data_align_center
            cell.fill = row_fill
        else:
            cell.alignment = data_align_left
            cell.fill = row_fill

    ws.row_dimensions[row_idx].height = 80

# --- ウィンドウ枠の固定 ---
ws.freeze_panes = "B2"

# --- 凡例シート作成 ---
ws2 = wb.create_sheet("凡例・スコア説明")
legend_data = [
    ["【スコア説明】", ""],
    ["", ""],
    ["IT未導入スコア（1〜5）", ""],
    ["5", "現在も手書き・FAX管理が継続中であることが公開情報で確認できる"],
    ["4", "導入前にアナログ管理だったことが事例・口コミで確認できる"],
    ["3", "間接的な証拠（業界特性・規模・地域性）からIT化が遅れている可能性が高い"],
    ["2", "一部IT化が進んでいるが、特定業務でアナログ管理が残っている"],
    ["1", "IT化が進んでいる（参考情報として掲載）"],
    ["", ""],
    ["誤配送・取り間違いリスクスコア（1〜5）", ""],
    ["5", "誤配送・取り間違いのトラブルが口コミ・フォーラム等で具体的に報告されている"],
    ["4", "業態・管理体制から誤配送リスクが高いことが事例で示唆されている"],
    ["3", "アナログ管理による情報連携ミスのリスクが推測される"],
    ["2", "リスクは低いが完全に排除されていない"],
    ["1", "誤配送リスクが低い（参考情報として掲載）"],
    ["", ""],
    ["【優先度】", ""],
    ["高（赤背景）", "総合スコア8以上、または直接的な証拠が複数ある企業"],
    ["中（黄背景）", "総合スコア6〜7、または間接的な証拠がある企業"],
    ["", ""],
    ["【調査方法】", ""],
    ["", "IT導入事例サイト（アセンド/ロジックス、Tradiss、神戸DXお助け隊等）、"],
    ["", "口コミサイト（エン転職、転職会議）、求人票、M&A事例記事、"],
    ["", "国土交通省・中小企業庁の公開資料をもとに調査・スコアリング"],
    ["", ""],
    ["【注意事項】", ""],
    ["", "本リストは公開情報をもとに作成したものです。"],
    ["", "IT導入状況は変化する場合があります（特に「導入事例」として掲載されている企業は現在改善中）。"],
    ["", "アプローチ時は最新情報を確認の上、ご活用ください。"],
]

ws2_font = Font(name="Meiryo UI", size=10)
ws2_bold = Font(name="Meiryo UI", size=10, bold=True)
ws2_header_fill = PatternFill("solid", fgColor="1F3864")
ws2_header_font = Font(name="Meiryo UI", size=11, bold=True, color="FFFFFF")

for row_idx, (col_a, col_b) in enumerate(legend_data, start=1):
    cell_a = ws2.cell(row=row_idx, column=1, value=col_a)
    cell_b = ws2.cell(row=row_idx, column=2, value=col_b)
    cell_a.alignment = Alignment(wrap_text=True, vertical="top")
    cell_b.alignment = Alignment(wrap_text=True, vertical="top")

    if col_a.startswith("【"):
        cell_a.font = Font(name="Meiryo UI", size=11, bold=True, color="1F3864")
        cell_b.font = ws2_font
    elif col_a in ("5", "4", "3", "2", "1"):
        cell_a.font = Font(name="Meiryo UI", size=10, bold=True)
        cell_a.fill = PatternFill("solid", fgColor=score_color(int(col_a)))
        cell_b.font = ws2_font
    elif col_a in ("高（赤背景）",):
        cell_a.fill = PatternFill("solid", fgColor=COLOR_HIGH_BG)
        cell_a.font = ws2_bold
        cell_b.font = ws2_font
    elif col_a in ("中（黄背景）",):
        cell_a.fill = PatternFill("solid", fgColor=COLOR_MID_BG)
        cell_a.font = ws2_bold
        cell_b.font = ws2_font
    else:
        cell_a.font = ws2_font
        cell_b.font = ws2_font

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 80

# --- 保存 ---
output_path = "/home/ubuntu/japan_transport_company_list.xlsx"
wb.save(output_path)
print(f"✅ Excelファイルを保存しました: {output_path}")
print(f"   - 企業数: {len(companies)}社")
print(f"   - シート: 運送会社リスト / 凡例・スコア説明")
