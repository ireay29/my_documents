#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
関西の運送・配送会社リスト作成スクリプト
観点：IT未導入・誤配送・取り間違い課題を抱えている会社
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== データ定義 =====
companies = [
    {
        "no": 1,
        "company_name": "株式会社池田商店",
        "location": "大阪府大阪市西成区北津守4-4-66",
        "prefecture": "大阪府",
        "business_type": "食肉の加工・販売・配送（学校・病院等への食品卸）",
        "employee_size": "不明",
        "vehicle_count": "不明",
        "website": "http://ikeda-syouten.com/",
        "it_lag_score": 5,
        "delivery_error_score": 5,
        "total_score": 10,
        "priority": "高",
        "it_lag_evidence": "「最も改善しなければならなかったのは、目視検品による誤出荷対策と賞味期限管理でした。」「総数38,000にも及ぶSKUを一つ一つ確認しながら…想像以上に大変でした。」「作業の標準化により属人化が完全に解消され」（WMS導入前は完全アナログ管理）",
        "delivery_error_evidence": "「かつて月間25〜30件あった誤出荷は、毎月減少してほぼ改善された」「特に食物アレルギーにかかわる食品は、間違って配送されれば命の危険にもつながる」（月25〜30件の誤出荷が公開されている）",
        "source": "https://www.logizard-zero.com/cases/ikeda-shoten.html",
        "notes": "WMS導入前は目視検品・アナログ管理。月間25〜30件の誤出荷が発生していたことが公開事例に明記されている。食物アレルギー対応食品も扱うため誤配送リスクが特に高い。",
    },
    {
        "no": 2,
        "company_name": "アサヒロジ株式会社（明石支店）",
        "location": "兵庫県明石市二見町南二見1-33",
        "prefecture": "兵庫県",
        "business_type": "食品物流（アサヒ飲料明石工場内の入出荷・配送管理）",
        "employee_size": "約2,000名（全社）",
        "vehicle_count": "不明",
        "website": "https://www.alogi.co.jp/",
        "it_lag_score": 5,
        "delivery_error_score": 3,
        "total_score": 8,
        "priority": "高",
        "it_lag_evidence": "「フォークリフトマンが回収した書類は事務所へ送られ、手作業で受付簿と照合して待機時間を割り出していました。」「1台のトラックが複数セットの伝票・作業指示書を持って複数の倉庫を回るため、各倉庫で回収された書類をトラック別に分別・照合するのがかなりの負担でした。」",
        "delivery_error_evidence": "「時刻の書き忘れ・書き間違いがあったり、正確さが担保できなかったりという問題もありました。」（手書き管理による記録ミスが発生していた）",
        "source": "https://www.hacobell.com/case_studies/pu1o5tlyoat",
        "notes": "2024年12月に「トラック簿」を導入し改善中。導入前は手書き伝票・手作業照合が常態化。複数倉庫・複数伝票の管理が属人化していた。",
    },
    {
        "no": 3,
        "company_name": "株式会社T.M.G",
        "location": "大阪府茨木市目垣1丁目23番15号",
        "prefecture": "大阪府",
        "business_type": "貨物軽自動車運送業・小口配送事業（Amazonデリバリープロバイダ）",
        "employee_size": "230名",
        "vehicle_count": "登録台数2,200台（グループ全体）",
        "website": "https://tmg-group.co.jp/",
        "it_lag_score": 2,
        "delivery_error_score": 5,
        "total_score": 7,
        "priority": "高",
        "it_lag_evidence": "独自の配送状況管理システムを導入済みとしているが、配送品質の管理が実態として追いついていない",
        "delivery_error_evidence": "Amazonセラーフォーラムに「不在票がない」「ポストに無理やり詰め込む（受け取りサイン必要なのに）」「配達時間指定が適当」「配完をあげておきながら商品を紛失した」「TMG様より詳細確認が困難である旨のご返答をいただきました」など具体的なトラブル報告が多数",
        "source": "https://sellercentral.amazon.co.jp/seller-forums/discussions/t/df4f1d782c476484ce3c2f0c22e9c728",
        "notes": "関西を中心に西日本に多くの拠点を持つAmazonデリバリープロバイダ。誤配送・紛失・破損のトラブル報告がフォーラムに多数掲載されている。",
    },
    {
        "no": 4,
        "company_name": "株式会社朋友",
        "location": "大阪府堺市中区平井848",
        "prefecture": "大阪府",
        "business_type": "一般貨物自動車運送・物流倉庫の人材派遣",
        "employee_size": "約120名",
        "vehicle_count": "約150台",
        "website": "https://foryou1992.co.jp/",
        "it_lag_score": 5,
        "delivery_error_score": 2,
        "total_score": 7,
        "priority": "高",
        "it_lag_evidence": "「デジタコのデータをエクセルに手打ちで転記をする作業が約100名分あり、毎日3時間ほどかかっていた」「パソコンが得意な社員とそうではない社員の間で業務処理の方法の違いが生じている」「昔ながらの管理法で逐一計算していた」（ロジックス導入前の状態）",
        "delivery_error_evidence": "誤配送の直接的な報告は見当たらないが、労務管理の属人化・アナログ管理体制から配送ミスのリスクが推測される",
        "source": "https://logix.ascendlogi.co.jp/case/foryou",
        "notes": "1992年創業。労基署からの是正勧告をきっかけにDX推進。ロジックス導入後は3時間の集計作業が10分に短縮。",
    },
    {
        "no": 5,
        "company_name": "株式会社南海エクスプレス",
        "location": "大阪府大阪市浪速区難波中1丁目10-4",
        "prefecture": "大阪府",
        "business_type": "航空貨物・海上貨物・通関業・ロジスティクス業",
        "employee_size": "191名",
        "vehicle_count": "不明",
        "website": "https://www.nankai-express.co.jp/",
        "it_lag_score": 4,
        "delivery_error_score": 4,
        "total_score": 8,
        "priority": "高",
        "it_lag_evidence": "「顧客企業から寄せられる配送依頼に対して、物流システムの担当者が手動で引当処理を実施していた」（Calsoft導入事例より）。OpenWork口コミでもIT化の遅れを示唆する記述あり",
        "delivery_error_evidence": "「その状況下だったからかミスが立て続けに起き、一時社内で問題とされました。」（就活会議口コミより）",
        "source": "https://www.calsoft.com/ja/導入事例/導入事例：南海エクスプレス-社のcargowise/",
        "notes": "大阪難波に本社を置く物流会社。手動引当処理が常態化しており、ミスが発生したことが口コミで報告されている。",
    },
    {
        "no": 6,
        "company_name": "藤原運輸株式会社",
        "location": "大阪府大阪市西区本田4-7-18",
        "prefecture": "大阪府",
        "business_type": "港湾運送業・自動車運送事業・倉庫業・通関業・船舶代理店業",
        "employee_size": "773名",
        "vehicle_count": "不明",
        "website": "https://fujiwaraunyu.com/",
        "it_lag_score": 4,
        "delivery_error_score": 1,
        "total_score": 5,
        "priority": "中",
        "it_lag_evidence": "「物流業界特有の紙文化が根強く、請求書や経理伝票、経費精算の処理が紙ベースで行われており、業務効率化が必要だった」（マネーフォワード クラウドインボイス導入事例より）",
        "delivery_error_evidence": "誤配送に関する直接的な記述は見当たらない",
        "source": "https://biz.moneyforward.com/case/12260/",
        "notes": "大阪港に拠点を持つ総合物流会社。請求書・経理伝票の紙管理が常態化していた。",
    },
    {
        "no": 7,
        "company_name": "新栄運輸株式会社",
        "location": "大阪府大阪市（野田駅周辺）",
        "prefecture": "大阪府",
        "business_type": "生鮮食品等の運送・送り状作成・伝票整理",
        "employee_size": "不明",
        "vehicle_count": "不明",
        "website": "不明",
        "it_lag_score": 5,
        "delivery_error_score": 3,
        "total_score": 8,
        "priority": "高",
        "it_lag_evidence": "「夜間に全国から届く生鮮食品等の送り状作成と伝票整理を手書きで行う事務作業です。」（求人情報より。現在も手書き伝票が継続中）",
        "delivery_error_evidence": "生鮮食品の多品種・夜間配送で手書き伝票管理のため、取り間違い・誤配送リスクが高い",
        "source": "https://xn--pckua2a7gp15o89zb.com/伝票整理-バイトの仕事",
        "notes": "求人情報から現在も手書きで送り状作成・伝票整理を行っていることが確認できた。夜間の生鮮食品配送という業態から誤配送リスクが高いと推測される。",
    },
    {
        "no": 8,
        "company_name": "株式会社オヅロジ",
        "location": "奈良県橿原市雲梯町104-1",
        "prefecture": "奈良県",
        "business_type": "一般貨物自動車運送・トラック輸送",
        "employee_size": "約40名",
        "vehicle_count": "不明",
        "website": "https://odulogi.biz/",
        "it_lag_score": 4,
        "delivery_error_score": 2,
        "total_score": 6,
        "priority": "中",
        "it_lag_evidence": "「非効率なアナログ業務（紙の不在票、手動でのルート作成など）に時間を取られるという構造的な課題」を解決しDXを推進中（Indeed求人情報より。以前はアナログ業務が中心だった）",
        "delivery_error_evidence": "手動ルート作成・紙の不在票管理による配送ミスリスクが存在していた",
        "source": "https://jp.indeed.com/q-奈良-運送会社-アナログ-求人.html",
        "notes": "奈良県橿原市の中小運送会社。DX推進中だが、以前はアナログ業務が中心だったことが求人情報から確認できる。",
    },
    {
        "no": 9,
        "company_name": "株式会社近藤製麺工場",
        "location": "京都府京都市南区上鳥羽苗代町",
        "prefecture": "京都府",
        "business_type": "製麺業・中華めん卸・麺のルート配送",
        "employee_size": "不明",
        "vehicle_count": "不明",
        "website": "https://www.kondoseimen.com/",
        "it_lag_score": 5,
        "delivery_error_score": 3,
        "total_score": 8,
        "priority": "高",
        "it_lag_evidence": "「手書き伝票や発注書の記入があるため日本語能力必須（N1相当）」と求人情報に明記。公式サイトにFAX番号の記載あり、支払いは現金のみ",
        "delivery_error_evidence": "手書き伝票・発注書による多品種配送のため、取り間違い・誤配送リスクが高い",
        "source": "https://www.hatalike.jp/viewjob/e074e11ad4b7e75d/",
        "notes": "京都市南区の製麺会社。自社でルート配送を行っており、手書き伝票が現在も使用されていることが求人情報から確認できた。",
    },
    {
        "no": 10,
        "company_name": "フジトランスポート株式会社",
        "location": "奈良県奈良市北之庄町723-13",
        "prefecture": "奈良県",
        "business_type": "一般貨物自動車運送・総合物流サービス",
        "employee_size": "3,043名（単体）/ 3,739名（グループ）",
        "vehicle_count": "2,780台（単体）/ 3,430台（グループ）",
        "website": "https://www.fujitransport.com/",
        "it_lag_score": 3,
        "delivery_error_score": 3,
        "total_score": 6,
        "priority": "中",
        "it_lag_evidence": "「電話、FAX、配車システムを使用し、トラック空車情報、お客様の荷物情報をマッチングさせるお仕事です。」（関西配車センターの求人情報より。FAXが現在も使用されている）",
        "delivery_error_evidence": "グループ会社・株式会社フジプラスにて「宛名データの不備により誤送付が発生」（情報管理体制の甘さを示唆）",
        "source": "https://fujitransport-recruit.net/jobfind-pc/job/All/7062",
        "notes": "奈良県に本社を置く大手物流会社。配車センターではFAXが現在も使用されている。グループ会社での誤送付事案あり。",
    },
]

# ===== Excel作成 =====
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "関西_運送会社リスト"

# --- カラー定義 ---
COLOR_HEADER_BG = "1F3864"
COLOR_HEADER_FG = "FFFFFF"
COLOR_HIGH_BG   = "FFE0E0"   # 薄赤（優先度：高）
COLOR_MID_BG    = "FFF3CD"   # 薄黄（優先度：中）
COLOR_SCORE_5   = "C6EFCE"
COLOR_SCORE_4   = "FFEB9C"
COLOR_SCORE_3   = "FFCC99"
COLOR_SCORE_2   = "FFC7CE"
COLOR_SCORE_1   = "FF9999"

# 都道府県カラー
PREF_COLORS = {
    "大阪府": "E8F4FD",
    "兵庫県": "FEF9E7",
    "京都府": "FDECEA",
    "滋賀県": "E8F8F5",
    "奈良県": "F5EEF8",
    "和歌山県": "FEF5E7",
}

def score_color(score):
    if score >= 5: return COLOR_SCORE_5
    elif score == 4: return COLOR_SCORE_4
    elif score == 3: return COLOR_SCORE_3
    elif score == 2: return COLOR_SCORE_2
    else: return COLOR_SCORE_1

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# --- ヘッダー定義 ---
headers = [
    ("No.", 5),
    ("会社名", 28),
    ("所在地", 32),
    ("都道府県", 10),
    ("業種・事業内容", 30),
    ("従業員数", 12),
    ("車両台数", 12),
    ("IT未導入\nスコア\n(1-5)", 10),
    ("誤配送\nリスク\nスコア(1-5)", 10),
    ("総合\nスコア", 8),
    ("優先度", 8),
    ("IT未導入の証拠・根拠", 55),
    ("誤配送・取り間違いの証拠・根拠", 55),
    ("情報ソース", 40),
    ("備考", 40),
    ("公式サイト", 35),
]

header_font = Font(name="Meiryo UI", bold=True, color=COLOR_HEADER_FG, size=9)
header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col_idx, (header_text, col_width) in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header_text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = col_width

ws.row_dimensions[1].height = 45

# --- データ行書き込み ---
data_font = Font(name="Meiryo UI", size=9)
data_align_center = Alignment(horizontal="center", vertical="top", wrap_text=True)
data_align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

for row_idx, c in enumerate(companies, start=2):
    priority = c["priority"]
    if priority == "高":
        row_bg = COLOR_HIGH_BG
    else:
        row_bg = COLOR_MID_BG

    row_fill = PatternFill("solid", fgColor=row_bg)

    row_data = [
        c["no"],
        c["company_name"],
        c["location"],
        c["prefecture"],
        c["business_type"],
        c["employee_size"],
        c["vehicle_count"],
        c["it_lag_score"],
        c["delivery_error_score"],
        c["total_score"],
        c["priority"],
        c["it_lag_evidence"],
        c["delivery_error_evidence"],
        c["source"],
        c["notes"],
        c["website"],
    ]

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.border = thin_border

        if col_idx in (8, 9, 10):
            cell.alignment = data_align_center
            cell.fill = PatternFill("solid", fgColor=score_color(value))
        elif col_idx in (1, 4, 11):
            cell.alignment = data_align_center
            cell.fill = row_fill
        else:
            cell.alignment = data_align_left
            cell.fill = row_fill

    ws.row_dimensions[row_idx].height = 90

# ウィンドウ枠固定
ws.freeze_panes = "B2"

# --- 凡例シート ---
ws2 = wb.create_sheet("凡例・スコア説明")
legend_data = [
    ["【スコア説明】", ""],
    ["", ""],
    ["IT未導入スコア（1〜5）", ""],
    ["5", "現在も手書き・FAX管理が継続中であることが公開情報（求人票・事例）で確認できる"],
    ["4", "導入前にアナログ管理だったことが事例・口コミで確認できる"],
    ["3", "間接的な証拠（業界特性・規模・地域性・FAX使用）からIT化が遅れている可能性が高い"],
    ["2", "一部IT化が進んでいるが、特定業務でアナログ管理が残っている"],
    ["1", "IT化が進んでいる（誤配送課題のみで掲載）"],
    ["", ""],
    ["誤配送・取り間違いリスクスコア（1〜5）", ""],
    ["5", "誤配送・取り間違いのトラブルが口コミ・フォーラム等で具体的に報告されている（件数明記含む）"],
    ["4", "業態・管理体制から誤配送リスクが高いことが事例・口コミで示唆されている"],
    ["3", "アナログ管理による情報連携ミスのリスクが推測される（業態・品種の多さから）"],
    ["2", "リスクは低いが完全に排除されていない"],
    ["1", "誤配送リスクが低い（参考情報として掲載）"],
    ["", ""],
    ["【優先度】", ""],
    ["高（赤背景）", "総合スコア7以上、または直接的な証拠が複数ある企業"],
    ["中（黄背景）", "総合スコア5〜6、または間接的な証拠がある企業"],
    ["", ""],
    ["【調査方法】", ""],
    ["", "IT導入事例サイト（アセンド/ロジックス、hacobell、logizard-zero等）、"],
    ["", "口コミサイト（エン転職、転職会議、OpenWork、就活会議）、"],
    ["", "求人票（Indeed、はたらいく、バイトル等）、"],
    ["", "Amazonセラーセントラルフォーラム、企業公式サイトをもとに調査・スコアリング"],
    ["", ""],
    ["【注意事項】", ""],
    ["", "本リストは公開情報をもとに作成したものです。"],
    ["", "IT導入状況は変化する場合があります（特に「導入事例」として掲載されている企業は現在改善中）。"],
    ["", "アプローチ時は最新情報を確認の上、ご活用ください。"],
]

ws2_font = Font(name="Meiryo UI", size=10)
ws2_bold = Font(name="Meiryo UI", size=10, bold=True)

for row_idx, (col_a, col_b) in enumerate(legend_data, start=1):
    cell_a = ws2.cell(row=row_idx, column=1, value=col_a)
    cell_b = ws2.cell(row=row_idx, column=2, value=col_b)
    cell_a.alignment = Alignment(wrap_text=True, vertical="top")
    cell_b.alignment = Alignment(wrap_text=True, vertical="top")

    if col_a.startswith("【"):
        cell_a.font = Font(name="Meiryo UI", size=11, bold=True, color="1F3864")
        cell_b.font = ws2_font
    elif col_a in ("5", "4", "3", "2", "1"):
        cell_a.font = Font(name="Meiryo UI", size=10, bold=True)
        cell_a.fill = PatternFill("solid", fgColor=score_color(int(col_a)))
        cell_b.font = ws2_font
    elif col_a in ("高（赤背景）",):
        cell_a.fill = PatternFill("solid", fgColor=COLOR_HIGH_BG)
        cell_a.font = ws2_bold
        cell_b.font = ws2_font
    elif col_a in ("中（黄背景）",):
        cell_a.fill = PatternFill("solid", fgColor=COLOR_MID_BG)
        cell_a.font = ws2_bold
        cell_b.font = ws2_font
    else:
        cell_a.font = ws2_font
        cell_b.font = ws2_font

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 80

# --- 保存 ---
output_path = "/home/ubuntu/kansai_transport_company_list.xlsx"
wb.save(output_path)
print(f"✅ Excelファイルを保存しました: {output_path}")
print(f"   - 企業数: {len(companies)}社")
print(f"   - 優先度「高」: {sum(1 for c in companies if c['priority'] == '高')}社")
print(f"   - 優先度「中」: {sum(1 for c in companies if c['priority'] == '中')}社")
